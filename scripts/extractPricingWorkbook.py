from pathlib import Path
import json
from openpyxl import load_workbook

workbook_path = Path("/home/ubuntu/upload/ZAVERRE_مقارنة_أسعار_81_سيارة_بالعربي.xlsx")
output_path = Path("/home/ubuntu/zafir-restore-v2/price-workbook-extract.json")
workbook = load_workbook(workbook_path, data_only=True)

summary_sheet = workbook["الملخص التنفيذي"]
comparison_sheet = workbook["مقارنة أسعار الـ81 سيارة"]

summary = [
    [cell.value for cell in row]
    for row in summary_sheet.iter_rows(values_only=False)
    if any(cell.value not in (None, "") for cell in row)
]
headers = [cell.value for cell in next(comparison_sheet.iter_rows(min_row=1, max_row=1, values_only=False))]
rows = []
for row in comparison_sheet.iter_rows(min_row=2, values_only=True):
    if not any(value not in (None, "") for value in row):
        continue
    rows.append(dict(zip(headers, row)))

output_path.write_text(json.dumps({
    "workbook": workbook_path.name,
    "sheets": workbook.sheetnames,
    "summary": summary,
    "comparisonHeaders": headers,
    "comparisonRows": rows,
}, ensure_ascii=False, indent=2), encoding="utf-8")

print(f"Extracted {len(rows)} comparison rows to {output_path}")
